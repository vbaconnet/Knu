"""
Informations
============

    * Fichier : ``Knu/Scripts/Librairies/outilsExcel.py``
    * Auteur: Victor Baconnet
    * Date de dernière modification: 27 août 2021

Description
===========

Outils d'écriture et d'édition de données et graphiques sur excel

Utilisation
============

Ces fonctions sont à importer directement dans vos programmes.

L'exécution de ce fichier ne produira aucun résultat.

Exemple d'utilisation
======================

Ecriture de données et affichage::
    
    import outilsExcel as oxc
    from openpyxl import Workbook
    from openpyxl.chart import ScatterChart
    from numpy import sin, pi, linspace
    from pandas import DataFrame
    
    # Génération de données
    t = linspace(0,10,1001)
    y1 = sin(2*pi*t)
    y2 = sin(2*pi*t/4)
    y3 = y1*y2
    
    # Tout mettre dans un dataframe
    df = DataFrame(data = [t,y1,y2,y3]).transpose() 
    df.columns = ["t","Periode 1s","Periode 4s","Produit"]
    
    # Initialisation du fichier excel et des feuilles de donées
    wkbk = Workbook()  # Objet worbook = fichier excel
    donnees = wkbk.active
    donnees.title = "Données"
    graphes = wkbk.create_sheet("Graphiques")
    
    # Creation des objets graphiques
    chart1 = ScatterChart(auto_axis = False)
    chart2 = ScatterChart(auto_axis = False)
    
    # Ecriture des données
    oxc.write(df, donnees, with_index = False)
    
    # Affichage des graphiques
    oxc.plot(chart1, [(1,2),(1,3)], donnees, graphes, "A1",
             xlabel = "Temps (s)", y_min_row = 2, x_min_row = 2,
             get_title = True, xlim=(0,10))
    
    oxc.plot(chart2, [(1,4)], donnees, graphes, "M1",
             xlabel = "Temps (s)", y_min_row = 2, x_min_row = 2,
             get_title = True, xlim=(0,10))
    
    #Sauvegarde
    wkbk.save("exemple.xlsx")
    

Fonctions
=========

"""


def write(data, sheet, with_index = False):
    """Ajoute (append) des données d'un DataFrame dans une feuille excel
    
    :param data: Tableau de données de type dataframe
    :type data: pandas.core.frame.DataFrame
    :param sheet: Feuille dans laquelle écrire
    :type sheet: openpyxl.worksheet.worksheet.Worksheet
    """
    
    from openpyxl.utils.dataframe import dataframe_to_rows

    for row in dataframe_to_rows(data, index = with_index):
        sheet.append(row)

def make_plot(chart, data_sheet, x_col, y_col, x_min_row = 2, y_min_row = 2,
    get_title = True):
    """Tracer un graphique sur une feuille à partir de données d'une autre 
    feuille. Trace colonne par colonne.

    :param chart: Objet LineChart sur lequel tracer
    :type chart: openpyxl.chart.line_chart.LineChart
    :param plot_sheet: Feuille sur laquelle tracer
    :type plot_sheet: openpyxl.worksheet.worksheet.Worksheet
    :param qf_sheet: Feuille à partir de laquelle récupérer les données
    :type qf_sheet: openpyxl.worksheet.worksheet.Worksheet
    :param x_col: Indice de la colonne des x (commence à 1)
    :type x_col: int
    :param y_col: Indice de la colonne des y (commence à 1)
    :type y_col: int
    :param get_title: Indique si il faut récupérer le titre de la colonne y ou non
    :type get_title: bool
    """
    
    print(f"Nouveau tracé  (x,y) : colonnes ({x_col}, {y_col})")
    
    from openpyxl.chart import Reference, Series

    x = Reference(worksheet = data_sheet,
                  min_row = 2,
                  min_col = x_col,
                  max_row = data_sheet.max_row,
                  max_col = x_col) 

    if get_title:
        y = Reference(worksheet = data_sheet,
                        min_row = y_min_row - 1,
                        min_col = y_col,
                        max_row = data_sheet.max_row,
                        max_col = y_col)
    else:
        y = Reference(worksheet = data_sheet,
                        min_row = y_min_row,
                        min_col = y_col,
                        max_row = data_sheet.max_row,
                        max_col = y_col)

    serie = Series(values = y,
                xvalues = x,
                title_from_data=get_title)

    chart.series.append(serie)


def plot(chart, cols, data_sheet, plot_sheet, where, title = "",
         xlabel = "", ylabel = "", y_min_row = 1, x_min_row = 2, height = 15,
         width = 20, legend_position = "t", display_legend = True,
         xlim = None, ylim = None, get_title = True, additional_info = ""):
    """Tracer plusieurs jeux de données sur un même graphique

    :param chart: Object chart sur lequel faire le graphique. Se crée avec la
      commande : ``chart = ScatterChart(auto_axis = False)``
    :type chart: ScatterChart
    :param cols: Colonnes à tracer. Syntaxe: ``[(x1,y1), (x2,y2), ...]``,
      où les ``xi`` et ``yi`` sont des entiers correspondant aux colonnes à tracer.
      La colonne A est la numéro 1, et ainsi de suite.
    :type cols: list(tuple(int))
    :param data_sheet: Feuille dans laquelle récupérer les données
    :type data_sheet: openpyxl.worksheet.worksheet.Worksheet
    :param plot_sheet: Feuille sur laquelle afficher le graphique
    :type plot_sheet: openpyxl.worksheet.worksheet.Worksheet
    :param where: Cellule où faire le graphique (coin supérieur gauche).
      Exemple "A1"
    :type where: str
    :param title: Titre du graphique
    :type title: str
    :param xlabel: Nom de l'axe x
    :type xlabel: str
    :param ylabel: Nom de l'axe y
    :type ylabel: str
    :param y_min_row: Ligne minimale à partir de laquelle lire les données y,
        AVEC le titre! Exemple si titre en A1 et données A2:100, ``y_min_row=1``
    :type y_min_row: int
    :param x_min_row: Ligne minimale à partir de laquelle lire les données x,
        SANS le titre! Exemple si titre en A1 et données A2:100, ``x_min_row=2``
    :type x_min_row: int
    :param height: hauteur du graphique
    :type height: int
    :param width: largeur du graphique
    :type width: int
    :param legend_position: Position de la légende. 

      * Top : ``t``
      * Bottom : ``b``
      * Left : ``l``
      * Right : ``r``
    """
    
    from openpyxl.chart.text import RichText
    import openpyxl.drawing.text as odt

    # Tracer chaque graphique
    for to_plot in cols:
        make_plot(chart, data_sheet, to_plot[0], to_plot[1], get_title = get_title)
    
    # Hauteur et largeur du graphe
    chart.height = height
    chart.width = width
    
    # Titre des axes
    chart.x_axis.title = xlabel
    chart.y_axis.title = ylabel
    
    chart.title = title

    # Fixer les limites des axes
    if ylim is not None:
        chart.y_axis.scaling.min = ylim[0]
        chart.y_axis.scaling.max = ylim[1] 
    
    if xlim is not None:
        chart.x_axis.scaling.min = xlim[0]
        chart.x_axis.scaling.max = xlim[1] 
    
    # Police d'écriture
    font = odt.Font(typeface="Calibri")
    sizeTitle = 1800 # Taille d'écriture pour les titres (taille 18)
    sizeOther = 1600 # Taille d'écriture pour les graduations et légende
    
    # Création d'objets CharacterProperties
    cpTitle = odt.CharacterProperties(latin=font, sz=sizeTitle, b=True) 
    cpOther = odt.CharacterProperties(latin=font, sz=sizeOther, b=False) 
    
    # Création d'objets ParagraphProperties
    ppTitle = odt.ParagraphProperties(defRPr=cpTitle)
    ppOther = odt.ParagraphProperties(defRPr=cpOther)
    
    # Création d'un objet RichText 
    rtpOther = RichText(p=[odt.Paragraph(pPr=ppOther, endParaRPr=cpOther)])
    
    # La police des graduations d'axes et de la légende se modifie à partir 
    # d'objets RichText
    chart.x_axis.txPr = rtpOther # Graduations axe x
    chart.y_axis.txPr = rtpOther # Graduations axe y
        
    # Positionner la légende
    if display_legend:
        chart.legend.position = legend_position
        chart.legend.txPr = rtpOther # Légende
    else:
        chart.legend = None
    
    # La police du titre des axes se modifie à partir d'objets ParagraphProperties
    chart.x_axis.title.tx.rich.p[0].pPr = ppTitle # Titre axe x
    chart.y_axis.title.tx.rich.p[0].pPr = ppTitle # Titre axe y

    # Ajouter le graphique à la feuille en position A1
    plot_sheet.add_chart(chart, where)